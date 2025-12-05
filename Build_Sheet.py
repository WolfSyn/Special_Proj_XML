# DO NOT COPYWRITE OR SELL THIS PRODUCT. 
# Created 10:40am 
# Author: Carlos Garcia

import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# ====== CONFIG ======
FOLDER = r"C:\\Users\\BigZ1\\OneDrive\\Desktop\\Test_XML"
OUTPUT_XLSX = os.path.join(FOLDER, "Sheet3_Builder_XML_Output.xlsx")

# Tests you want in Sheet3
TESTS_TO_INCLUDE = ["Max Power (dBm)", "Throughput (%)"]

# Optional: narrow to specific channels per band (exact match)
# Example: CHANNELS_KEEP = {"n77": ["650000/650000"], "B12": None, "B5": None}
CHANNELS_KEEP = None  # leave None to keep all channels

# ====== XML READER ======
# We know rows are under <TestSteps>/<TestStep>
XML_ROW_PATH = ".//TestStep"  # [1](https://imperoelectronics-my.sharepoint.com/personal/cgarcia_encorerepair_com/Documents/Microsoft%20Copilot%20Chat%20Files/P_354240388124327_D=11-21-2025_T=17-33-56.xml)

def read_one_xml(path: str) -> pd.DataFrame:
    # Parse XML rows
    df_raw = pd.read_xml(path, xpath=XML_ROW_PATH)
    # Keep only fields we need for Sheet3
    cols_needed = ['Serial_No','Band','UL_DL_Chans','Test','MeasValue']
    # Basic sanity check in case of slight schema differences
    for c in cols_needed:
        if c not in df_raw.columns:
            raise ValueError(f"{os.path.basename(path)} missing column '{c}'. Present: {df_raw.columns.tolist()}")
    df = df_raw[cols_needed].copy()
    return df

def normalize_and_filter(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [c.strip() for c in df.columns]
    df = df[df['Test'].isin(TESTS_TO_INCLUDE)].copy()
    if CHANNELS_KEEP:
        keep_rows = []
        for _, row in df.iterrows():
            b = str(row['Band']); ch = str(row['UL_DL_Chans'])
            if b in CHANNELS_KEEP:
                allowed = CHANNELS_KEEP[b]
                keep_rows.append((allowed is None) or (ch in allowed))
            else:
                keep_rows.append(True)
        df = df[pd.Series(keep_rows, index=df.index)]
    return df

def make_sheet3(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame({
        'Serial_No': df['Serial_No'],
        'Band': df['Band'],
        'UL_DL_Chans': df['UL_DL_Chans'],
        'Test': df['Test'],
        'Original Reading': pd.NA,             # keep blank (change to df['MeasValue'] if you want)
        "GroupO's Reading": df['MeasValue'],
        'Applies': pd.NA,
        'RF1': pd.NA,
        'RF6': pd.NA,
        'RF12': pd.NA
    }).sort_values(['Serial_No','Band','UL_DL_Chans','Test']).reset_index(drop=True)
    return out

def write_with_conditional_formatting(df: pd.DataFrame, xlsx_path: str):
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as wr:
        df.to_excel(wr, sheet_name='Sheet3', index=False)
    wb = load_workbook(xlsx_path)
    ws = wb['Sheet3']
    max_row = ws.max_row
    # Column F holds GroupO's Reading
    rng = f'F2:F{max_row}'
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    rule = CellIsRule(operator='lessThan', formula=['95'], stopIfTrue=True, fill=red_fill)
    ws.conditional_formatting.add(rng, rule)
    # Optional: column widths
    widths = {'A': 16, 'B': 8, 'C': 16, 'D': 18, 'E': 18, 'F': 18, 'G': 10, 'H': 8, 'I': 8, 'J': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    wb.save(xlsx_path)

def main():
    xml_files = glob.glob(os.path.join(FOLDER, "*.xml"))
    if not xml_files:
        raise RuntimeError("No XML files found in the folder.")

    frames = []
    for path in xml_files:
        try:
            df = read_one_xml(path)
            df = normalize_and_filter(df)
            frames.append(df)
        except Exception as e:
            print(f"[WARN] Skipped {os.path.basename(path)}: {e}")

    if not frames:
        raise RuntimeError("No parseable XML files. Check XML_ROW_PATH or column names.")

    raw = pd.concat(frames, ignore_index=True)
    sheet3 = make_sheet3(raw)
    write_with_conditional_formatting(sheet3, OUTPUT_XLSX)
    print(f"Done. Wrote {len(sheet3)} rows into: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
